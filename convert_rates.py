import json
import os
import glob
from datetime import datetime, timezone
import openpyxl

EXCEL_PATTERN = "Customer Rate Tariff Template_*.xlsx"
INSURANCE_AMOUNT = 200

DEST_MAP = {
    "TT": "TT",
    "GUY": "GUY",
    "GY": "GUY",
    "SUR": "SUR",
    "TRINIDAD EXPORTS": "Trinidad Exports",
    "PRINT FE-TT": "Print FE-TT",
    "COL": "COL",
}

DEST_LABELS = {
    "TT": "Trinidad",
    "GUY": "Guyana",
    "SUR": "Suriname",
    "Trinidad Exports": "Trinidad Exports",
    "Print FE-TT": "Far East to Trinidad",
    "COL": "Colombia",
}

# Canonical field names mapped from various header spellings
HEADER_MAP = {
    "pol":                   "pol",
    "pod":                   "pod",
    "of/bunker":             "of_bunker",
    "thc":                   "thc",
    "lac":                   "lac",
    "local charges":         "lac",
    "lac/local charges":     "lac",
    "isps":                  "isps",
    "container inspection":  "container_inspection",
    "other port charges":    "other_port_charges",
    "local handling":        "local_handling",
    "admin":                 "admin",
    "docs":                  "docs",
    "gri":                   "gri",
    "total without insurance":   "total",
    "total w/out insurance":     "total",
    "total w/o insurance":       "total",
    "insurance":             "insurance_col",
    "total with insurance":  "total_with_insurance",
    "transit time":          "transit_time",
    "validity":              "validity",
    "carrier":               "carrier",
    "comment":               "comment",
}

def find_excel_file():
    matches = sorted(glob.glob(EXCEL_PATTERN))
    if not matches:
        raise FileNotFoundError(f"No Excel file matching '{EXCEL_PATTERN}' found in {os.getcwd()}")
    return matches[-1]

def safe_float(value):
    if value is None:
        return None
    if isinstance(value, str) and value.strip().upper() in ('N/A', '-', '', 'NONE'):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

def safe_str(value):
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None

def build_cell_map(ws):
    merge_values = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(merge_range.min_row, merge_range.min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_values[(row, col)] = top_left
    cell_map = {}
    for row in ws.iter_rows():
        for cell in row:
            key = (cell.row, cell.column)
            cell_map[key] = merge_values.get(key, cell.value)
    return cell_map

def get(cell_map, r, c):
    return cell_map.get((r, c))

def is_section_title(val):
    if val and isinstance(val, str) and 'TARIFF' in val.upper():
        return val.strip()
    return None

def is_header_row(cell_map, r, max_col):
    for c in range(1, max_col + 1):
        val = safe_str(get(cell_map, r, c))
        if val and 'OF/BUNKER' in val.upper():
            return True
    return False

def parse_header_row(cell_map, r, max_col):
    col_map = {}
    for c in range(1, max_col + 1):
        val = safe_str(get(cell_map, r, c))
        if val:
            key = val.lower().strip()
            canonical = HEADER_MAP.get(key)
            if canonical and canonical not in col_map:
                col_map[canonical] = c
    return col_map

def extract_lane(title):
    return title.replace(' SHIPPING TARIFF', '').replace(' TARIFF', '').strip()

def parse_sheet(ws):
    cell_map = build_cell_map(ws)
    max_row = ws.max_row
    max_col = ws.max_column
    sections = {}
    current_lane = None
    current_col_map = {}

    for r in range(1, max_row + 1):
        pol_val = safe_str(get(cell_map, r, 2))

        # Check for section title
        title = is_section_title(pol_val)
        if title:
            current_lane = extract_lane(title)
            if current_lane not in sections:
                sections[current_lane] = []
            current_col_map = {}
            continue

        # Check for header row
        if is_header_row(cell_map, r, max_col):
            current_col_map = parse_header_row(cell_map, r, max_col)
            continue

        if current_lane is None or not current_col_map:
            continue

        size = safe_str(get(cell_map, r, 4))
        if not size or size.upper() not in ('20FT', '40FT', '20GP', '40GP', '20HC', '40HC'):
            continue

        pol = safe_str(get(cell_map, r, current_col_map.get('pol', 2)))
        pod = safe_str(get(cell_map, r, current_col_map.get('pod', 3)))
        if not pol or not pod:
            continue

        def gcol(field):
            c = current_col_map.get(field)
            return get(cell_map, r, c) if c else None

        validity_raw = gcol('validity')
        if hasattr(validity_raw, 'strftime'):
            validity = validity_raw.strftime('%d/%m/%Y')
        else:
            validity = safe_str(validity_raw)

        entry = {
            "pol":                   pol,
            "pod":                   pod,
            "size":                  size,
            "of_bunker":             safe_float(gcol('of_bunker')),
            "thc":                   safe_float(gcol('thc')),
            "lac":                   safe_float(gcol('lac')),
            "isps":                  safe_float(gcol('isps')),
            "container_inspection":  safe_float(gcol('container_inspection')),
            "other_port_charges":    safe_float(gcol('other_port_charges')),
            "local_handling":        safe_float(gcol('local_handling')),
            "admin":                 safe_float(gcol('admin')),
            "docs":                  safe_float(gcol('docs')),
            "gri":                   safe_float(gcol('gri')),
            "total":                 safe_float(gcol('total')),
            "total_with_insurance":  safe_float(gcol('total_with_insurance')),
            "insurance":             INSURANCE_AMOUNT,
            "transit_time":          safe_str(gcol('transit_time')),
            "validity":              validity,
            "carrier":               safe_str(gcol('carrier')),
            "comment":               safe_str(gcol('comment')),
        }
        sections[current_lane].append(entry)

    return sections

def escape_html(text):
    if not text:
        return ""
    return (str(text)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))

def generate_html(output):
    generated_at = output["generated_at"]
    source_file = output["source_file"]
    destinations = output["destinations"]

    rows = []
    rows.append("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Ramps Logistics FCL Rates</title>
<style>
body { font-family: Arial, sans-serif; padding: 20px; max-width: 1200px; margin: 0 auto; }
h1 { color: #1a1a1a; }
h2 { color: #1a1a1a; border-bottom: 2px solid #ccc; padding-bottom: 4px; margin-top: 40px; }
h3 { color: #333; margin-top: 20px; }
table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
th, td { border: 1px solid #ddd; padding: 6px 10px; text-align: left; font-size: 13px; }
th { background: #f5f5f5; font-weight: bold; }
tr:nth-child(even) { background: #fafafa; }
</style>
</head>
<body>
<h1>Ramps Logistics FCL Shipping Rates</h1>""")

    rows.append(f'<p><strong>Last updated:</strong> {escape_html(generated_at)} &nbsp;|&nbsp; <strong>Source:</strong> {escape_html(source_file)}</p>')

    dest_order = ["TT", "GUY", "SUR", "Trinidad Exports", "Print FE-TT", "COL"]
    for dest_key in dest_order:
        if dest_key not in destinations:
            continue
        label = DEST_LABELS.get(dest_key, dest_key)
        rows.append(f"<h2>{escape_html(label)}</h2>")
        lanes = destinations[dest_key]
        for lane, entries in lanes.items():
            if not entries:
                continue
            rows.append(f"<h3>{escape_html(lane)}</h3>")
            rows.append("<table>")
            rows.append("<tr><th>POL</th><th>POD</th><th>Size</th><th>Carrier</th><th>Total (with ins.)</th><th>Transit</th><th>Validity</th><th>Comment</th></tr>")
            for e in entries:
                total = e.get("total_with_insurance")
                total_str = f"USD {total}" if total is not None else ""
                transit = escape_html(e.get("transit_time") or "")
                comment = escape_html(e.get("comment") or "")
                rows.append(
                    f'<tr>'
                    f'<td>{escape_html(e["pol"])}</td>'
                    f'<td>{escape_html(e["pod"])}</td>'
                    f'<td>{escape_html(e["size"])}</td>'
                    f'<td>{escape_html(e.get("carrier") or "")}</td>'
                    f'<td>{escape_html(total_str)}</td>'
                    f'<td>{transit}</td>'
                    f'<td>{escape_html(e.get("validity") or "")}</td>'
                    f'<td>{comment}</td>'
                    f'</tr>'
                )
            rows.append("</table>")

    rows.append("</body>\n</html>")
    return "\n".join(rows)

def convert(excel_path):
    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    destinations = {}
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        dest_code = DEST_MAP.get(sheet_name.strip().upper(), sheet_name)
        sections = parse_sheet(ws)
        row_count = sum(len(v) for v in sections.values())
        print(f"  Sheet '{sheet_name}' -> {len(sections)} lanes | {row_count} rows")
        if row_count == 0:
            continue
        if dest_code not in destinations:
            destinations[dest_code] = {}
        destinations[dest_code].update(sections)
        total_rows += row_count

    output = {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_file": os.path.basename(excel_path),
        "destinations": destinations,
    }

    with open("rates.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    html_content = generate_html(output)
    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html_content)

    size_kb = round(os.path.getsize("rates.json") / 1024, 1)
    print(f"\nDone - {total_rows} rows -> rates.json ({size_kb} KB) + index.html regenerated")

if __name__ == "__main__":
    excel_file = find_excel_file()
    convert(excel_file)
